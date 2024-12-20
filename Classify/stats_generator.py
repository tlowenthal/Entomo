import pandas as pd
from openpyxl import Workbook
import os
from openpyxl.styles import PatternFill

prediction_file = 'result.csv' #path to the resulted file from predict.py
correction_file = 'correct_predictions.csv' #path to the correct predictions file

base_name = os.path.splitext(os.path.basename(prediction_file))[0] 
output_file = f'{base_name}_stat.xlsx'

pred_df = pd.read_csv(prediction_file, sep=',', header=0)
correct_df = pd.read_csv(correction_file, sep=';', header = 0)

pred_df = pred_df.sort_values(by='Image_Name').reset_index(drop=True)

result_df = pred_df.copy()
result_df = pd.merge(pred_df, correct_df[['Image_Name', 'Predicted_Species']], on='Image_Name', how='left', suffixes=('', '_Correct'))

result_df['Correct'] = result_df.apply(
    lambda row: row['Predicted_Species'] == row['Predicted_Species_Correct'],
    axis=1
)

VP = ((result_df['Correct'] == True) & (result_df['Predicted_Species'].isin(correct_df['Predicted_Species']))).sum()
VN = ((result_df['Correct'] == False) & (~result_df['Predicted_Species'].isin(correct_df['Predicted_Species']))).sum()
FP = ((result_df['Correct'] == False) & (result_df['Predicted_Species'].isin(correct_df['Predicted_Species']))).sum()
FN = ((result_df['Correct'] == True) & (~result_df['Predicted_Species'].isin(correct_df['Predicted_Species']))).sum()


accuracy = (VP + VN) / len(result_df)
rappel = VP / (VP + FN) if (VP + FN) > 0 else 0
precision = VP / (VP + FP) if (VP + FP) > 0 else 0
f1_score = (2 * precision * rappel / (precision + rappel)) if (precision + rappel) > 0 else 0
specificity = VN / (VN + FP) if (VN + FP) > 0 else 0
npv = VN / (VN + FN) if (VN + FN) > 0 else 0
nb_correct = VP
tot = len(result_df)


with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    result_df.to_excel(writer, index=False, sheet_name='Comparaison')
    workbook = writer.book
    worksheet = writer.sheets['Comparaison']

    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    for index, row in result_df.iterrows():
        if not row['Correct']:
            for col in range(len(row)):
                worksheet.cell(row=index + 2, column=col + 1).fill = red_fill


    stats_df = pd.DataFrame({
        'Metric': ['Accuracy', 'Recall', 'Precision', 'F1-Score', 'Specificity', 'Negative Predictive Value','Number of correct predictions','Total','VP','FP','VN', 'FN'],
        'Value': [accuracy, rappel, precision, f1_score, specificity, npv, nb_correct, tot, VP, FP, VN, FN]
    })

    stats_df.to_excel(writer, index=False, sheet_name='Statistiques')

print("New file saved")